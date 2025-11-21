"""
Report Generator - Creates Excel and PDF reports for kitchen staff
"""

from datetime import datetime
from typing import Dict
import json


class ReportGenerator:
    """Generates formatted reports (Excel, PDF, JSON)"""

    def __init__(self, report_data: Dict):
        """
        Initialize report generator

        Args:
            report_data: Report data from MealPlanner
        """
        self.report_data = report_data

    def generate_excel(self, filename: str):
        """
        Generate Excel report

        Args:
            filename: Output filename
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            print("Error: openpyxl not installed. Run: pip install openpyxl")
            return

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Meal Planning"

        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        total_font = Font(bold=True, size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = "KITCHEN MEAL PLANNING REPORT"
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Date range
        ws.merge_cells('A2:H2')
        date_cell = ws['A2']
        date_cell.value = f"Week of {self.report_data['start_date']} to {self.report_data['end_date']}"
        date_cell.alignment = Alignment(horizontal='center')

        # Generated timestamp
        ws.merge_cells('A3:H3')
        gen_cell = ws['A3']
        gen_cell.value = f"Generated: {self.report_data['generated_at']}"
        gen_cell.alignment = Alignment(horizontal='center')
        gen_cell.font = Font(italic=True, size=9)

        # Headers
        headers = ['Date', 'Day', 'Guests', 'Breakfast', 'Lunch', 'Dinner', 'Conference', 'Reservations']
        row = 5
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Data rows
        row = 6
        for day in self.report_data['daily_breakdown']:
            ws.cell(row=row, column=1, value=day['date']).border = border
            ws.cell(row=row, column=2, value=day['day_name']).border = border
            ws.cell(row=row, column=3, value=day['total_guests']).border = border
            ws.cell(row=row, column=4, value=day['breakfast']).border = border
            ws.cell(row=row, column=5, value=day['lunch']).border = border
            ws.cell(row=row, column=6, value=day['dinner']).border = border
            ws.cell(row=row, column=7, value=day['conference']).border = border
            ws.cell(row=row, column=8, value=day['reservation_count']).border = border

            # Center align numbers
            for col in range(3, 9):
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')

            row += 1

            # Add notes if available
            notes = day.get('notes', {})
            if notes:
                note_style = Font(italic=True, size=9, color="555555")
                note_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

                if notes.get('general'):
                    ws.merge_cells(f'A{row}:H{row}')
                    note_cell = ws.cell(row=row, column=1, value=f"  NOTE: {notes['general']}")
                    note_cell.font = note_style
                    note_cell.fill = note_fill
                    note_cell.alignment = Alignment(horizontal='left', vertical='center')
                    row += 1

                if notes.get('dietary'):
                    ws.merge_cells(f'A{row}:H{row}')
                    note_cell = ws.cell(row=row, column=1, value=f"  DIETARY: {notes['dietary']}")
                    note_cell.font = note_style
                    note_cell.fill = note_fill
                    note_cell.alignment = Alignment(horizontal='left', vertical='center')
                    row += 1

                if notes.get('location'):
                    ws.merge_cells(f'A{row}:H{row}')
                    note_cell = ws.cell(row=row, column=1, value=f"  LOCATION: {notes['location']}")
                    note_cell.font = note_style
                    note_cell.fill = note_fill
                    note_cell.alignment = Alignment(horizontal='left', vertical='center')
                    row += 1

                if notes.get('special'):
                    ws.merge_cells(f'A{row}:H{row}')
                    note_cell = ws.cell(row=row, column=1, value=f"  SPECIAL: {notes['special']}")
                    note_cell.font = note_style
                    note_cell.fill = note_fill
                    note_cell.alignment = Alignment(horizontal='left', vertical='center')
                    row += 1

        # Weekly totals
        totals = self.report_data['weekly_totals']
        ws.cell(row=row, column=1, value="WEEKLY TOTAL").font = total_font
        ws.cell(row=row, column=1).fill = total_fill
        ws.cell(row=row, column=1).border = border

        ws.cell(row=row, column=2, value="").fill = total_fill
        ws.cell(row=row, column=2).border = border

        ws.cell(row=row, column=3, value=totals['total_guests']).font = total_font
        ws.cell(row=row, column=3).fill = total_fill
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3).border = border

        ws.cell(row=row, column=4, value=totals['breakfast']).font = total_font
        ws.cell(row=row, column=4).fill = total_fill
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=4).border = border

        ws.cell(row=row, column=5, value=totals['lunch']).font = total_font
        ws.cell(row=row, column=5).fill = total_fill
        ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5).border = border

        ws.cell(row=row, column=6, value=totals['dinner']).font = total_font
        ws.cell(row=row, column=6).fill = total_fill
        ws.cell(row=row, column=6).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=6).border = border

        ws.cell(row=row, column=7, value=totals['conference']).font = total_font
        ws.cell(row=row, column=7).fill = total_fill
        ws.cell(row=row, column=7).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=7).border = border

        ws.cell(row=row, column=8, value="").fill = total_fill
        ws.cell(row=row, column=8).border = border

        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 14

        # Save
        wb.save(filename)
        print(f"\nExcel report saved: {filename}")

    def generate_pdf(self, filename: str):
        """
        Generate PDF report

        Args:
            filename: Output filename
        """
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib import colors
            from reportlab.lib.units import inch
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.enums import TA_CENTER
        except ImportError:
            print("Error: reportlab not installed. Run: pip install reportlab")
            return

        # Create PDF
        doc = SimpleDocTemplate(filename, pagesize=letter)
        elements = []

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#366092'),
            alignment=TA_CENTER,
            spaceAfter=12
        )

        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Normal'],
            fontSize=12,
            alignment=TA_CENTER,
            spaceAfter=6
        )

        # Title
        elements.append(Paragraph("KITCHEN MEAL PLANNING REPORT", title_style))
        elements.append(Paragraph(
            f"Week of {self.report_data['start_date']} to {self.report_data['end_date']}",
            subtitle_style
        ))
        elements.append(Paragraph(
            f"Generated: {self.report_data['generated_at']}",
            subtitle_style
        ))
        elements.append(Spacer(1, 0.3 * inch))

        # Table data
        data = [['Date', 'Day', 'Guests', 'Breakfast', 'Lunch', 'Dinner', 'Conference', 'Res.']]

        for day in self.report_data['daily_breakdown']:
            data.append([
                day['date'],
                day['day_name'],
                day['total_guests'],
                day['breakfast'],
                day['lunch'],
                day['dinner'],
                day['conference'],
                day['reservation_count']
            ])

            # Add notes rows
            notes = day.get('notes', {})
            if notes:
                if notes.get('general'):
                    data.append(['', f"NOTE: {notes['general']}", '', '', '', '', '', ''])
                if notes.get('dietary'):
                    data.append(['', f"DIETARY: {notes['dietary']}", '', '', '', '', '', ''])
                if notes.get('location'):
                    data.append(['', f"LOCATION: {notes['location']}", '', '', '', '', '', ''])
                if notes.get('special'):
                    data.append(['', f"SPECIAL: {notes['special']}", '', '', '', '', '', ''])

        # Weekly totals
        totals = self.report_data['weekly_totals']
        data.append([
            'TOTAL',
            '',
            totals['total_guests'],
            totals['breakfast'],
            totals['lunch'],
            totals['dinner'],
            totals['conference'],
            ''
        ])

        # Create table
        table = Table(data, colWidths=[1.1*inch, 0.9*inch, 0.7*inch, 0.9*inch, 0.7*inch, 0.7*inch, 0.9*inch, 0.6*inch])

        # Table style
        table.setStyle(TableStyle([
            # Header row
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),

            # Data rows
            ('ALIGN', (2, 1), (-1, -2), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -2), 10),
            ('GRID', (0, 0), (-1, -2), 1, colors.black),

            # Totals row
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 11),
            ('ALIGN', (0, -1), (-1, -1), 'CENTER'),
            ('GRID', (0, -1), (-1, -1), 1, colors.black),
        ]))

        elements.append(table)

        # Build PDF
        doc.build(elements)
        print(f"PDF report saved: {filename}")

    def generate_json(self, filename: str):
        """
        Generate JSON report

        Args:
            filename: Output filename
        """
        with open(filename, 'w') as f:
            json.dump(self.report_data, f, indent=2)
        print(f"JSON report saved: {filename}")
