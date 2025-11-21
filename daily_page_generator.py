"""
Daily Page Report Generator
Creates clean, one-page-per-day reports with timeline layout
Similar to Oracle system reports
"""

from datetime import datetime
from typing import Dict, List
import json


class DailyPageGenerator:
    """Generates clean daily page reports with timeline layout"""

    def __init__(self, report_data: Dict):
        self.report_data = report_data

    def generate_daily_pages_excel(self, filename: str):
        """
        Generate Excel workbook with one clean page per day

        Args:
            filename: Output filename
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            print("Error: openpyxl not installed. Run: pip install openpyxl")
            return

        wb = openpyxl.Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Create overview sheet
        self._create_overview_sheet(wb)

        # Create one sheet per day
        for day in self.report_data['daily_breakdown']:
            self._create_daily_page(wb, day)

        wb.save(filename)
        print(f"\nDaily pages report saved: {filename}")

    def _create_overview_sheet(self, wb):
        """Create week overview sheet"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        ws = wb.create_sheet("Week Overview", 0)

        # Title
        ws.merge_cells('A1:F1')
        title = ws['A1']
        title.value = "WEEKLY MEAL PLANNING OVERVIEW"
        title.font = Font(bold=True, size=16, color="FFFFFF")
        title.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # Week dates
        ws.merge_cells('A2:F2')
        week_info = ws['A2']
        week_info.value = f"Week of {self.report_data['start_date']} to {self.report_data['end_date']}"
        week_info.font = Font(size=12)
        week_info.alignment = Alignment(horizontal='center')

        # Headers
        headers = ['Date', 'Day', 'Guests', 'Breakfast', 'Lunch', 'Dinner']
        row = 4
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

        # Data rows
        row = 5
        for day in self.report_data['daily_breakdown']:
            ws.cell(row=row, column=1, value=day['date'])
            ws.cell(row=row, column=2, value=day['day_name'])
            ws.cell(row=row, column=3, value=day['total_guests']).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=4, value=day['breakfast']).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=5, value=day['lunch']).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=6, value=day['dinner']).alignment = Alignment(horizontal='center')

            # Add border
            for col in range(1, 7):
                ws.cell(row=row, column=col).border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12

    def _create_daily_page(self, wb, day):
        """
        Create one clean page for a single day
        Timeline format with events on left, details on right
        """
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        day_name = day['day_name']
        date = day['date']
        sheet_name = f"{day_name} {date}"

        ws = wb.create_sheet(sheet_name)

        # Set page setup for printing
        ws.page_setup.paperSize = ws.PAPERSIZE_LETTER
        ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
        ws.print_options.horizontalCentered = True

        # ============ HEADER SECTION ============
        row = 1

        # Day and Date Header
        ws.merge_cells(f'A{row}:F{row}')
        header = ws.cell(row=row, column=1)
        header.value = f"{day_name.upper()}"
        header.font = Font(bold=True, size=24, color="FFFFFF")
        header.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[row].height = 35
        row += 1

        # Date and Guest Count
        ws.merge_cells(f'A{row}:F{row}')
        subheader = ws.cell(row=row, column=1)
        subheader.value = f"{date}  •  {day['total_guests']} Guests  •  {day['breakfast']}B/{day['lunch']}L/{day['dinner']}D"
        subheader.font = Font(size=12, italic=True)
        subheader.fill = PatternFill(start_color="E8F0F8", end_color="E8F0F8", fill_type="solid")
        subheader.alignment = Alignment(horizontal='center')
        ws.row_dimensions[row].height = 25
        row += 1

        # Spacer
        row += 1

        notes = day.get('notes', {})

        # General overview if available
        if notes and notes.get('general'):
            ws.merge_cells(f'A{row}:F{row}')
            overview = ws.cell(row=row, column=1)
            overview.value = f"📋 {notes['general']}"
            overview.font = Font(size=11, bold=True, italic=True)
            overview.fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
            overview.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws.row_dimensions[row].height = 25
            row += 1

        # Dietary warnings if available
        if notes and notes.get('dietary'):
            ws.merge_cells(f'A{row}:F{row}')
            dietary = ws.cell(row=row, column=1)
            dietary.value = f"⚠️  DIETARY: {notes['dietary']}"
            dietary.font = Font(size=10, bold=True, color="CC0000")
            dietary.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
            dietary.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws.row_dimensions[row].height = 20
            row += 1

        row += 1

        # ============ TIMELINE SECTION ============
        # Column headers for timeline
        ws.cell(row=row, column=1, value="TIME").font = Font(bold=True, size=10)
        ws.cell(row=row, column=1).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='center')

        ws.cell(row=row, column=2, value="SERVICE / EVENT").font = Font(bold=True, size=10)
        ws.cell(row=row, column=2).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='left')

        ws.cell(row=row, column=3, value="GUEST GROUP").font = Font(bold=True, size=10)
        ws.cell(row=row, column=3).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        ws.cell(row=row, column=3).alignment = Alignment(horizontal='left')

        ws.merge_cells(f'D{row}:F{row}')
        ws.cell(row=row, column=4, value="DETAILS").font = Font(bold=True, size=10)
        ws.cell(row=row, column=4).fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        ws.cell(row=row, column=4).alignment = Alignment(horizontal='left')

        row += 1

        # Build timeline entries
        timeline = self._build_timeline(notes)

        # Add timeline entries
        for entry in timeline:
            # Time column
            time_cell = ws.cell(row=row, column=1, value=entry['time'])
            time_cell.font = Font(bold=True, size=11)
            time_cell.alignment = Alignment(horizontal='center', vertical='top')
            time_cell.border = Border(left=Side(style='thin'), right=Side(style='medium'))

            # Service/Event column
            service_cell = ws.cell(row=row, column=2, value=entry['service'])
            service_cell.font = Font(size=10, bold=True if entry.get('highlight') else False)
            service_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            # Guest Group column
            group_cell = ws.cell(row=row, column=3, value=entry.get('guest_group', ''))
            group_cell.font = Font(size=9, italic=True)
            group_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            # Details column
            ws.merge_cells(f'D{row}:F{row}')
            details_cell = ws.cell(row=row, column=4, value=entry['details'])
            details_cell.font = Font(size=9)
            details_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            # Highlight important entries
            if entry.get('highlight'):
                service_cell.fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
                group_cell.fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
                details_cell.fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")

            # Set row height based on content
            ws.row_dimensions[row].height = max(30, len(entry['details']) / 3)

            row += 1

        # ============ BOTTOM SECTION ============
        row += 1

        # Preparation checklist if available
        if notes and notes.get('ingredients_prep'):
            ws.merge_cells(f'A{row}:F{row}')
            prep_header = ws.cell(row=row, column=1, value="PREPARATION CHECKLIST")
            prep_header.font = Font(bold=True, size=11)
            prep_header.fill = PatternFill(start_color="E8F0F8", end_color="E8F0F8", fill_type="solid")
            row += 1

            for item in notes['ingredients_prep']:
                ws.merge_cells(f'A{row}:F{row}')
                prep_item = ws.cell(row=row, column=1, value=f"☐ {item}")
                prep_item.font = Font(size=9)
                prep_item.alignment = Alignment(horizontal='left', wrap_text=True)
                row += 1

        # Column widths
        ws.column_dimensions['A'].width = 8   # Time
        ws.column_dimensions['B'].width = 15  # Service/Event
        ws.column_dimensions['C'].width = 18  # Guest Group
        ws.column_dimensions['D'].width = 18  # Details
        ws.column_dimensions['E'].width = 18  # Details cont
        ws.column_dimensions['F'].width = 18  # Details cont

    def _build_timeline(self, notes: Dict) -> List[Dict]:
        """
        Build chronological timeline from notes

        Returns:
            List of timeline entries with time, service, and details
        """
        timeline = []

        if not notes:
            return [
                {'time': '07:00', 'service': 'Breakfast Service', 'guest_group': '', 'details': 'Regular service', 'highlight': False},
                {'time': '12:00', 'service': 'Lunch Service', 'guest_group': '', 'details': 'Regular service', 'highlight': False},
                {'time': '19:00', 'service': 'Dinner Service', 'guest_group': '', 'details': 'Regular service', 'highlight': False}
            ]

        # Extract breakfast times
        if notes.get('breakfast') and notes['breakfast'].get('service_times'):
            for service in notes['breakfast']['service_times']:
                details = f"{service.get('covers', '?')} covers\n"
                details += f"Location: {service.get('location', 'TBD')}\n"
                if service.get('notes'):
                    details += service['notes']
                if service.get('menu'):
                    details += f"\n{service['menu']}"

                timeline.append({
                    'time': service.get('time', ''),
                    'service': 'BREAKFAST',
                    'guest_group': service.get('group', service.get('notes', '').split('-')[0].strip() if service.get('notes') else ''),
                    'details': details,
                    'highlight': service.get('covers', 0) > 50
                })

        # Extract coffee breaks
        if notes.get('coffee_breaks'):
            for break_info in notes['coffee_breaks']:
                details = f"{break_info.get('covers', '?')} attendees\n"
                details += f"Location: {break_info.get('location', '')}\n"
                details += f"Items: {break_info.get('items', '')}"

                timeline.append({
                    'time': break_info.get('time', ''),
                    'service': 'Coffee Break',
                    'guest_group': break_info.get('group', ''),
                    'details': details,
                    'highlight': False
                })

        # Extract lunch times
        if notes.get('lunch') and notes['lunch'].get('service_times'):
            for service in notes['lunch']['service_times']:
                details = f"{service.get('covers', '?')} covers\n"
                details += f"Location: {service.get('location', 'TBD')}\n"
                if service.get('notes'):
                    details += service['notes'] + "\n"
                if service.get('menu'):
                    details += service['menu']
                if service.get('timing'):
                    details += f"\n⏰ {service['timing']}"

                timeline.append({
                    'time': service.get('time', ''),
                    'service': 'LUNCH',
                    'guest_group': service.get('group', service.get('notes', '').split('-')[0].strip() if service.get('notes') else ''),
                    'details': details,
                    'highlight': service.get('covers', 0) > 50
                })

        # Extract dinner times
        if notes.get('dinner') and notes['dinner'].get('service_times'):
            for service in notes['dinner']['service_times']:
                details = f"{service.get('covers', '?')} covers\n"
                details += f"Location: {service.get('location', 'TBD')}\n"
                if service.get('notes'):
                    details += service['notes'] + "\n"
                if service.get('menu'):
                    details += service['menu']
                if service.get('timing'):
                    details += f"\n⏰ {service['timing']}"

                timeline.append({
                    'time': service.get('time', ''),
                    'service': 'DINNER',
                    'guest_group': service.get('group', service.get('notes', '').split('-')[0].strip() if service.get('notes') else ''),
                    'details': details,
                    'highlight': service.get('covers', 0) > 50 or 'wedding' in service.get('notes', '').lower()
                })

        # Extract timeline events if available
        if notes.get('timeline'):
            for event in notes['timeline']:
                if ' - ' in event:
                    time_part, event_part = event.split(' - ', 1)
                    timeline.append({
                        'time': time_part.strip(),
                        'service': 'Event',
                        'guest_group': '',
                        'details': event_part.strip(),
                        'highlight': False
                    })

        # Sort by time
        timeline.sort(key=lambda x: x['time'])

        # If empty, add default entries
        if not timeline:
            timeline = [
                {'time': '07:00', 'service': 'Breakfast Service', 'guest_group': '', 'details': 'Regular service', 'highlight': False},
                {'time': '12:00', 'service': 'Lunch Service', 'guest_group': '', 'details': 'Regular service', 'highlight': False},
                {'time': '19:00', 'service': 'Dinner Service', 'guest_group': '', 'details': 'Regular service', 'highlight': False}
            ]

        return timeline
