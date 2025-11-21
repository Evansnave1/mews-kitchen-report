"""
Detailed Report Generator - Enhanced kitchen planning reports
Includes service times, covers per sitting, ingredients prep, staffing
"""

from datetime import datetime
from typing import Dict
import json


class DetailedReportGenerator:
    """Generates detailed kitchen planning reports with service times and preparation details"""

    def __init__(self, report_data: Dict):
        """
        Initialize detailed report generator

        Args:
            report_data: Report data from MealPlanner
        """
        self.report_data = report_data

    def generate_detailed_excel(self, filename: str):
        """
        Generate detailed Excel report with service times and preparation info

        Args:
            filename: Output filename
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            print("Error: openpyxl not installed. Run: pip install openpyxl")
            return

        wb = openpyxl.Workbook()

        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Create sheets
        self._create_summary_sheet(wb)
        self._create_daily_details_sheets(wb)
        self._create_prep_sheet(wb)
        self._create_staffing_sheet(wb)

        wb.save(filename)
        print(f"\nDetailed Excel report saved: {filename}")

    def _create_summary_sheet(self, wb):
        """Create summary overview sheet"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        ws = wb.create_sheet("Weekly Summary", 0)

        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        ws.merge_cells('A1:I1')
        title = ws['A1']
        title.value = "WEEKLY MEAL PLANNING - DETAILED SUMMARY"
        title.font = Font(bold=True, size=16)
        title.alignment = Alignment(horizontal='center')

        ws.merge_cells('A2:I2')
        date_cell = ws['A2']
        date_cell.value = f"Week of {self.report_data['start_date']} to {self.report_data['end_date']}"
        date_cell.alignment = Alignment(horizontal='center')

        # Headers
        headers = ['Date', 'Day', 'Guests', 'Breakfast', 'Lunch', 'Dinner', 'Conf', 'Room Service', 'Notes']
        row = 4
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        # Data
        row = 5
        for day in self.report_data['daily_breakdown']:
            notes = day.get('notes', {})

            ws.cell(row=row, column=1, value=day['date']).border = border
            ws.cell(row=row, column=2, value=day['day_name']).border = border
            ws.cell(row=row, column=3, value=day['total_guests']).border = border
            ws.cell(row=row, column=4, value=day['breakfast']).border = border
            ws.cell(row=row, column=5, value=day['lunch']).border = border
            ws.cell(row=row, column=6, value=day['dinner']).border = border
            ws.cell(row=row, column=7, value=day['conference']).border = border

            # Room service count
            rs_count = notes.get('room_service', {}).get('breakfast_orders', 0) + \
                      notes.get('room_service', {}).get('lunch_orders', 0) + \
                      notes.get('room_service', {}).get('dinner_orders', 0) if isinstance(notes.get('room_service'), dict) else 0
            ws.cell(row=row, column=8, value=rs_count).border = border

            # Summary notes
            summary_notes = notes.get('general', '') if notes else ''
            cell = ws.cell(row=row, column=9, value=summary_notes)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True, vertical='top')

            # Center align numbers
            for col in range(3, 9):
                ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')

            row += 1

        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 8
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 40

    def _create_daily_details_sheets(self, wb):
        """Create detailed sheet for each day"""
        for day in self.report_data['daily_breakdown']:
            date = day['date']
            day_name = day['day_name']
            sheet_name = f"{day_name[:3]} {date[-5:]}"  # e.g., "Mon 01-15"

            notes = day.get('notes', {})
            if not notes or (isinstance(notes, dict) and not any([notes.get('breakfast'), notes.get('lunch'), notes.get('dinner')])):
                continue  # Skip days without detailed notes

            ws = wb.create_sheet(sheet_name)
            self._populate_daily_sheet(ws, day, notes)

    def _populate_daily_sheet(self, ws, day, notes):
        """Populate a daily details sheet"""
        from openpyxl.styles import Font, PatternFill, Alignment

        # Styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        section_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        section_font = Font(bold=True, size=12)

        row = 1

        # Title
        ws.merge_cells(f'A{row}:E{row}')
        title = ws.cell(row=row, column=1)
        title.value = f"{day['day_name']} {day['date']} - DETAILED PLAN"
        title.font = Font(bold=True, size=14)
        title.alignment = Alignment(horizontal='center')
        row += 2

        # General notes
        if notes.get('general'):
            ws.merge_cells(f'A{row}:E{row}')
            cell = ws.cell(row=row, column=1, value=f"OVERVIEW: {notes['general']}")
            cell.font = Font(bold=True, italic=True)
            row += 1

        if notes.get('dietary'):
            ws.merge_cells(f'A{row}:E{row}')
            cell = ws.cell(row=row, column=1, value=f"⚠️  DIETARY: {notes['dietary']}")
            cell.font = Font(bold=True, color="FF0000")
            row += 1

        row += 1

        # Breakfast section
        if notes.get('breakfast'):
            row = self._add_meal_section(ws, row, "BREAKFAST", notes['breakfast'], section_fill, section_font, header_fill, header_font)

        # Lunch section
        if notes.get('lunch'):
            row = self._add_meal_section(ws, row, "LUNCH", notes['lunch'], section_fill, section_font, header_fill, header_font)

        # Dinner section
        if notes.get('dinner'):
            row = self._add_meal_section(ws, row, "DINNER", notes['dinner'], section_fill, section_font, header_fill, header_font)

        # Coffee breaks
        if notes.get('coffee_breaks'):
            row = self._add_coffee_breaks(ws, row, notes['coffee_breaks'], section_fill, section_font)

        # Room service
        if notes.get('room_service'):
            row = self._add_room_service(ws, row, notes['room_service'], section_fill, section_font)

        # Ingredients prep
        if notes.get('ingredients_prep'):
            row = self._add_prep_list(ws, row, "INGREDIENTS TO PREP", notes['ingredients_prep'], section_fill, section_font)

        # Timeline
        if notes.get('timeline'):
            row = self._add_timeline(ws, row, notes['timeline'], section_fill, section_font)

        # Column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 40

    def _add_meal_section(self, ws, row, meal_name, meal_data, section_fill, section_font, header_fill, header_font):
        """Add a meal service section"""
        from openpyxl.styles import Alignment

        # Section header
        ws.merge_cells(f'A{row}:E{row}')
        header = ws.cell(row=row, column=1, value=meal_name)
        header.fill = section_fill
        header.font = section_font
        row += 1

        # Service times table
        if meal_data.get('service_times'):
            headers = ['Time', 'Covers', 'Location', 'Menu/Notes']
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            row += 1

            for service in meal_data['service_times']:
                ws.cell(row=row, column=1, value=service.get('time', ''))
                ws.cell(row=row, column=2, value=service.get('covers', ''))
                ws.cell(row=row, column=3, value=service.get('location', ''))

                notes_text = service.get('notes', '')
                if service.get('menu'):
                    notes_text += f"\n{service['menu']}"
                if service.get('timing'):
                    notes_text += f"\n⏰ {service['timing']}"

                cell = ws.cell(row=row, column=4, value=notes_text)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                ws.row_dimensions[row].height = max(15 * notes_text.count('\n'), 20)
                row += 1

        # Special requests
        if meal_data.get('special_requests'):
            row += 1
            ws.merge_cells(f'A{row}:E{row}')
            cell = ws.cell(row=row, column=1, value="SPECIAL REQUESTS:")
            cell.font = Font(bold=True, italic=True)
            row += 1

            for request in meal_data['special_requests']:
                ws.merge_cells(f'A{row}:E{row}')
                cell = ws.cell(row=row, column=1, value=f"  • {request}")
                cell.alignment = Alignment(wrap_text=True)
                row += 1

        row += 1
        return row

    def _add_coffee_breaks(self, ws, row, breaks, section_fill, section_font):
        """Add coffee breaks section"""
        ws.merge_cells(f'A{row}:E{row}')
        header = ws.cell(row=row, column=1, value="COFFEE BREAKS")
        header.fill = section_fill
        header.font = section_font
        row += 1

        for break_info in breaks:
            ws.cell(row=row, column=1, value=break_info.get('time', ''))
            ws.cell(row=row, column=2, value=f"{break_info.get('covers', '')} pax")
            ws.cell(row=row, column=3, value=break_info.get('location', ''))
            ws.cell(row=row, column=4, value=break_info.get('items', ''))
            row += 1

        row += 1
        return row

    def _add_room_service(self, ws, row, rs_data, section_fill, section_font):
        """Add room service section"""
        ws.merge_cells(f'A{row}:E{row}')
        header = ws.cell(row=row, column=1, value="ROOM SERVICE")
        header.fill = section_fill
        header.font = section_font
        row += 1

        ws.cell(row=row, column=1, value="Breakfast orders:")
        ws.cell(row=row, column=2, value=rs_data.get('breakfast_orders', 0))
        row += 1

        ws.cell(row=row, column=1, value="Lunch orders:")
        ws.cell(row=row, column=2, value=rs_data.get('lunch_orders', 0))
        row += 1

        ws.cell(row=row, column=1, value="Dinner orders:")
        ws.cell(row=row, column=2, value=rs_data.get('dinner_orders', 0))
        row += 1

        if rs_data.get('notes'):
            ws.merge_cells(f'A{row}:E{row}')
            ws.cell(row=row, column=1, value=f"Note: {rs_data['notes']}")
            row += 1

        row += 1
        return row

    def _add_prep_list(self, ws, row, title, items, section_fill, section_font):
        """Add preparation list"""
        ws.merge_cells(f'A{row}:E{row}')
        header = ws.cell(row=row, column=1, value=title)
        header.fill = section_fill
        header.font = section_font
        row += 1

        for item in items:
            ws.merge_cells(f'A{row}:E{row}')
            cell = ws.cell(row=row, column=1, value=f"  ☐ {item}")
            row += 1

        row += 1
        return row

    def _add_timeline(self, ws, row, timeline, section_fill, section_font):
        """Add event timeline"""
        ws.merge_cells(f'A{row}:E{row}')
        header = ws.cell(row=row, column=1, value="EVENT TIMELINE")
        header.fill = section_fill
        header.font = section_font
        row += 1

        for event in timeline:
            ws.merge_cells(f'A{row}:E{row}')
            cell = ws.cell(row=row, column=1, value=f"  {event}")
            row += 1

        row += 1
        return row

    def _create_prep_sheet(self, wb):
        """Create ingredients prep summary sheet"""
        from openpyxl.styles import Font, Alignment

        ws = wb.create_sheet("Prep Checklist")

        # Title
        ws.merge_cells('A1:D1')
        title = ws['A1']
        title.value = "WEEKLY INGREDIENTS PREP CHECKLIST"
        title.font = Font(bold=True, size=14)
        title.alignment = Alignment(horizontal='center')

        row = 3
        for day in self.report_data['daily_breakdown']:
            notes = day.get('notes', {})
            if notes and notes.get('ingredients_prep'):
                ws.merge_cells(f'A{row}:D{row}')
                header = ws.cell(row=row, column=1, value=f"{day['day_name']} {day['date']}")
                header.font = Font(bold=True, size=12)
                row += 1

                for item in notes['ingredients_prep']:
                    ws.cell(row=row, column=1, value="☐")
                    ws.merge_cells(f'B{row}:D{row}')
                    ws.cell(row=row, column=2, value=item)
                    row += 1

                row += 1

        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 50

    def _create_staffing_sheet(self, wb):
        """Create staffing requirements sheet"""
        from openpyxl.styles import Font, Alignment

        ws = wb.create_sheet("Staffing")

        # Title
        ws.merge_cells('A1:E1')
        title = ws['A1']
        title.value = "WEEKLY STAFFING REQUIREMENTS"
        title.font = Font(bold=True, size=14)
        title.alignment = Alignment(horizontal='center')

        # Headers
        headers = ['Date', 'Day', 'Breakfast', 'Lunch', 'Dinner', 'Notes']
        row = 3
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = Font(bold=True)
        row += 1

        for day in self.report_data['daily_breakdown']:
            notes = day.get('notes', {})
            staffing = notes.get('staffing', {}) if notes else {}

            if staffing:
                ws.cell(row=row, column=1, value=day['date'])
                ws.cell(row=row, column=2, value=day['day_name'])
                ws.cell(row=row, column=3, value=staffing.get('breakfast_staff', ''))
                ws.cell(row=row, column=4, value=staffing.get('lunch_staff', ''))
                ws.cell(row=row, column=5, value=staffing.get('dinner_staff', ''))

                cell = ws.cell(row=row, column=6, value=staffing.get('notes', ''))
                cell.alignment = Alignment(wrap_text=True)
                row += 1

        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 15
        ws.column_dimensions['F'].width = 40
